# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                      # Для .xlsx
import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, quoted, dump_cell, currencyType, subInParentheses
import csv



def getXlsString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена со скидкой','цена_') :
            if getCell(row=i, col=j, isDigit='N', sheet=sh) == '' :       # .find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCell(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCell(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена','цена1') :
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('Звоните') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert2csv( fileNameIn, cfgFName ):
    basicNamelist, basic = config_read( cfgFName, 'basic' )
    csvFName  = basic['filename_out']
    sheetName = basic['sheetname']
    
    log.debug('Reading price ' + fileNameIn )
    book = openpyxl.load_workbook(filename = fileNameIn, read_only=False, keep_vba=False, data_only=False)  # xlsx
    #sheet = book.worksheets[0]                                                                             # xlsx
    sheet = book[sheetName]                                                                                 # xlsx 
    log.info('-------------------  '+sheet.title +'  ----------')                                           # xlsx
    sheetNames = book.get_sheet_names()                                                                     # xlsx

#    book = xlrd.open_workbook( fileNameIn.encode('cp1251'), formatting_info=True)                       # xls
#    sheet = book.sheets()[0]                                                                            # xls
#    log.info('-------------------  '+sheet.name +'  ----------')                                        # xls

    out_cols, out_template = config_read(cfgFName, 'cols_out')
    in_cols,  in_cols_j    = config_read(cfgFName, 'cols_in')
    #brands,   discount     = config_read(cfgFName, 'discount')
    for k in in_cols_j.keys():
        p = in_cols_j[k].find(' ')
        if p>0 :
            in_cols_j[k] = int(in_cols_j[k][ :p])                                   # -1              # xls
        else:
            in_cols_j[k] = int(in_cols_j[k]     )                                   # -1              # xls
    #for k in discount.keys():
    #    discount[k] = (100 - int(discount[k]))/100
    #print(discount)

    outFile = open( csvFName, 'w', newline='', encoding='CP1251', errors='replace')
    csvWriter = csv.DictWriter(outFile, fieldnames=out_cols )
    csvWriter.writeheader()

    '''                                            # Блок проверки свойств для распознавания групп      XLSX                                  
    for i in range(2393, 2397):                                                         
        i_last = i
        ccc = sheet.cell( row=i, column=in_cols_j['группа'] )
        print(i, ccc.value)
        print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, ccc.font.color.rgb, '------', ccc.fill.fgColor.rgb)
        print('------')
    '''
    '''                                            # Блок проверки свойств для распознавания групп      XLS                                  
    for i in range(0, 75):                                                         
        xfx = sheet.cell_xf_index(i, 0)
        xf  = book.xf_list[xfx]
        bgci  = xf.background.pattern_colour_index
        fonti = xf.font_index
        ccc = sheet.cell(i, 0)
        if ccc.value == None :
            print (i, colSGrp, 'Пусто!!!')
            continue
                                         # Атрибуты шрифта для настройки конфига
        font = book.font_list[fonti]
        print( '---------------------- Строка', i, '-----------------------', sheet.cell(i, 0).value)
        print( 'background_colour_index=',bgci)
        print( 'fonti=', fonti, '           xf.alignment.indent_level=', xf.alignment.indent_level)
        print( 'bold=', font.bold)
        print( 'weight=', font.weight)
        print( 'height=', font.height)
        print( 'italic=', font.italic)
        print( 'colour_index=', font.colour_index )
        print( 'name=', font.name)
    return
    '''

    ssss    = []
    brand   = ''
    grp     = ''
    subgrp  = ''
    brand_koeft = 1
    recOut  ={}

#   for i in range(1, sheet.nrows) :                                    # xls
    for i in range(1, sheet.max_row +1) :                               # xlsx
        i_last = i
        try:
            #print('i =',i,)
            '''                                                         # xls 
            xfx = sheet.cell_xf_index(i, 0)
            xf  = book.xf_list[xfx]
            level = xf.alignment.indent_level
            bgci  = xf.background.pattern_colour_index
            ccc   = sheet.cell(i, 0)
            value = ccc.value   
            '''
            impValues = getXlsxString(sheet, i, in_cols_j)
                
            if impValues['код'] == '' or impValues['код'] == 'Арт.' :  # Пустая строка
                print (i, 'Пусто!!!')
                continue
            else :                                                     # Обычная строка
                for outColName in out_template.keys() :
                    if (outColName == 'закупка') and (impValues['бренд'].lower() == 'epiphan') and ( float( impValues['цена1']) >10000 ) :
                        shablon = str( float(impValues['цена1']) * 0.7  )
                    elif (outColName == 'закупка') :
                        shablon = str( float(impValues['цена1']) * 0.75 )
                    else:
                        shablon = out_template[outColName]
                        for key in impValues.keys():
                            if shablon.find(key) >= 0 :
                                shablon = shablon.replace(key, impValues[key])
                    recOut[outColName] = shablon
                csvWriter.writerow(recOut)

        except Exception as e:
            print(e)
            if str(e) == "'NoneType' object has no attribute 'rgb'":
                pass
            else:
                log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'.' )

    log.info('Обработано ' +str(i_last)+ ' строк.')
    outFile.close()



def config_read( cfgFName, partName ):
    log.debug('Reading config ' + cfgFName )
    config = configparser.ConfigParser()
    keyList = []
    keyDict = {}
    if  os.path.exists(cfgFName):     
        config.read( cfgFName, encoding='utf-8')
        keyList = config.options(partName)
        for vName in keyList :
            if ('' != config.get(partName, vName)) :
                keyDict[vName] = config.get(partName, vName)
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    
    return keyList, keyDict



def download( dealerName ):
    pathDwnld = './tmp'
    pathPython2 = 'c:/Python27/python.exe'
    retCode = False
    fUnitName = os.path.join( dealerName +'_unittest.py')
    if  not os.path.exists(fUnitName):
        log.debug( 'Отсутствует юниттест для загрузки прайса ' + fUnitName)
    else:
        dir_befo_download = set(os.listdir(pathDwnld))
        os.system( pathPython2 + ' ' + fUnitName)                                                           # Вызов unittest'a
        dir_afte_download = set(os.listdir(pathDwnld))
        new_files = list( dir_afte_download.difference(dir_befo_download))
        if len(new_files) == 1 :   
            new_file = new_files[0]                                                     # загружен ровно один файл. 
            new_ext  = os.path.splitext(new_file)[-1]
            DnewFile = os.path.join( pathDwnld,new_file)
            new_file_date = os.path.getmtime(DnewFile)
            log.info( 'Скачанный файл ' +DnewFile + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) ) )
            if new_ext == '.zip':                                                       # Архив. Обработка не завершена
                log.debug( 'Zip-архив. Разархивируем.')
                work_dir = os.getcwd()                                                  
                os.chdir( os.path.join( pathDwnld ))
                dir_befo_download = set(os.listdir(os.getcwd()))
                os.system('unzip -oj ' + new_file)
                os.remove(new_file)   
                dir_afte_download = set(os.listdir(os.getcwd()))
                new_files = list( dir_afte_download.difference(dir_befo_download))
                if len(new_files) == 1 :   
                    new_file = new_files[0]                                             # разархивирован ровно один файл. 
                    new_ext  = os.path.splitext(new_file)[-1]
                    DnewFile = os.path.join( os.getcwd(),new_file)
                    new_file_date = os.path.getmtime(DnewFile)
                    log.debug( 'Файл из архива ' +DnewFile + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) )     )
                    DnewPrice = DnewFile
                elif len(new_files) >1 :
                    log.debug( 'В архиве не единственный файл. Надо разбираться.')
                    DnewPrice = "dummy"
                else:
                    log.debug( 'Нет новых файлов после разархивации. Загляни в папку юниттеста поставщика.')
                    DnewPrice = "dummy"
                os.chdir(work_dir)
            elif new_ext not in ( '.csv', '.htm', '.xls', '.xlsx', 'xlsb'):
                DnewPrice = DnewFile                                             # Имя скачанного прайса
            if DnewPrice != "dummy" :
                FoldName = 'old_' + dealerName + new_ext                         # Старая копия прайса, для сравнения даты
                FnewName = 'new_' + dealerName + new_ext                         # Предыдущий прайс, с которым работает макрос
                if  (not os.path.exists( FnewName)) or new_file_date > os.path.getmtime(FnewName) : 
                    log.debug( 'Предыдущего прайса нет или скачанный файл новее. Копируем его.' )
                    if os.path.exists( FoldName): os.remove( FoldName)
                    if os.path.exists( FnewName): os.rename( FnewName, FoldName)
                    shutil.copy2(DnewPrice, FnewName)
                    retCode = True
                else:
                    log.debug( 'Файл у поставщика старый, копироавать его не надо.' )
                # Убрать скачанные файлы
                if  os.path.exists(DnewPrice):  os.remove(DnewPrice)   
            
        elif len(new_files) == 0 :        
            log.debug( 'Не удалось скачать файл прайса ')
        else:
            log.debug( 'Скачалось несколько файлов. Надо разбираться ...')
    return FnewName,retCode



def is_file_expiry(priceName, cfgFName):
    basicNamelist, basic = config_read( cfgFName, 'basic' )
    qty_days = basic['срок годности']
    p = qty_days.find(' ')
    if p>0 :
        qty_days = int(qty_days[ :p]) 
    else:
        qty_days = int(qty_days     )
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( priceName):
        price_datetime = os.path.getmtime(priceName)
    else:
        log.error('Не найден файл  '+ priceName)
        return False

    if price_datetime+qty_seconds < time.time() :
        log.error('Файл устарел!  Допустимый период '+ str(qty_days)+' дней.')
        return False
    else:
        return True



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def main( dealerName):
    make_loger()
    log.info('         '+dealerName )
    csvFName   = ('csv_'+dealerName+'.csv').lower()
    cfgFName   = ('cfg_'+dealerName+'.cfg').lower()
    priceName  = ('new_'+dealerName+'.xlsx').lower()
    
    fileName, result = download(dealerName)
    #fileName, result = 'new_brullov.xlsb',True
    if result:
        if  is_file_expiry( fileName, cfgFName):
            os.system( dealerName + '_converter_xlsx.xlsm')
            convert2csv( priceName, cfgFName)
    if os.path.exists( csvFName    ) : shutil.copy2( csvFName ,    'c://AV_PROM/prices/' + dealerName +'/'+csvFName )
    if os.path.exists( 'python.log') : shutil.copy2( 'python.log', 'c://AV_PROM/prices/' + dealerName +'/python.log')
    if os.path.exists( 'python.1'  ) : shutil.copy2( 'python.log', 'c://AV_PROM/prices/' + dealerName +'/python.1'  )


if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
